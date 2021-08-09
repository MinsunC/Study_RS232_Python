import openpyxl
import numpy as np

# Log에서 남기고 싶은 Parameter

dic = {
        "GroupCycle_AO": 0,
        "StepNo_AO": 0,
        # "Gauge2PrsAI": 0,
        # "Gauge4PrsAI": 0,
        # "T3BP_Set_Pressure": 0,
        "T3BP_Pressure": 0,
        # "T3BP_Set_Position": 0,
        # "T3BP_Position": 0,
        # "MFC01StPtAO": 0,
        # "MFC02StPtAO": 0,
        # "MFC03StPtAO": 0,
        # "MFC04StPtAO": 0,
        # "MFC01FlwAI": 0,
        # "MFC02FlwAI": 0,
        # "MFC03FlwAI": 0,
        # "MFC04FlwAI": 0,
        # "HT06_TmpSpAIO": 0,
        # "HT06_TmpRdAI": 0,
        # "HT07_TmpSpAIO": 0,
        # "HT07_TmpRdAI": 0,
        # "HT08_TmpSpAIO": 0,
        # "HT08_TmpRdAI": 0,
        # "HT06_OutPwCurtAI": 0,
        # "HT07_OutPwCurtAI": 0,
        # "HT08_OutPwCurtAI": 0
        # "T3BP_SetGain": 0,
        # "T3BP_SetPhase": 0
}

filename = input("""파일명(xlsx 확장자 포함 파일 이름): """)

workbook = openpyxl.load_workbook(filename)

sheet1 = workbook.worksheets[0]  # 기존 Data
sheet2 = workbook.create_sheet()  # Filtering한 Data 넣을 sheet

stepTime = 0
currentStep = 0
buffer = []
buffer2 = []
b = [0, 0, 0]
c = [0]
data = np.zeros(201)
logInterval = 0.2
# for row in sheet1.iter_rows():
#     if i == 1:
#         j = 0
#         for cell in row:
#             if cell.value in dic:
#                 dic[cell.value] = j
#             j += 1
#         i = 0
#         continue
#     if cycle != row[dic["GroupCycle_AO"]].value:
#         cycle = row[dic["GroupCycle_AO"]].value
#         a.append([row[dic["GroupCycle_AO"]].value])

skip = 1
m = 0
for row in sheet1.iter_rows():
    if skip == 1:
        for i in range(len(row)):
            buffer.append(b)
            buffer2.append(row[i].value)
        sheet2.append(buffer2)
        i = 0
        continue

    if currentStep == row[0].value:
        stepTime += 1
    elif currentStep != row[0].value:
        currentStep = row[0].value
        stepTime = 0

    # Stable time
    if currentStep % 2 == 0:
        for i in range(len(row)-1):
            buffer[i][stepTime % 3] = row[i+1].value
        for i in range(len(row)-1):
            j = 0
            for s in buffer[i]:
                if np.abs(s - currentStep/2) <= 0.003:
                    j += 1
                if j == len(buffer[i]):
                    data[i+1] = stepTime * logInterval
        m = 1
    elif currentStep % 2 == 1 and m == 1:
        data[0] = currentStep/2
        sheet2.append(data)
        data = np.zeros(201)

        m = 0

workbook.remove(workbook.worksheets[0])
workbook.save("filtered_" + filename)
