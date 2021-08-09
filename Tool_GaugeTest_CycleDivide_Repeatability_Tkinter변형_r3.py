import openpyxl
import numpy
from tkinter import *
from tkinter import filedialog
from os.path import basename

loglists = 0
window = 0
window2 = 0
columnlists = []


class LogFilter:
    global loglists

    def __init__(self):
        global window
        global columnlists

        window = Tk()
        window.geometry("640x400+200+200")
        window.title("Log Filter")

        self.button1 = Button(window, text="Selection: Log files", command=lambda: self.logfiles())
        self.button1.pack(pady=10)

        self.button2 = Button(window, text="Selection: Save directory", command=lambda: self.savedir())
        self.button2.pack(pady=10)

        self.button3 = Button(window, text="Select column", command=lambda: self.selectcolumn())
        self.button3.pack(pady=20)

        self.button4 = Button(window, text="Start data filtering", command=lambda: self.startfilter(columnlists))
        self.button4.pack(pady=20)

        window.mainloop()

    def logfiles(self):
        global loglists

        loglists = filedialog.askopenfilenames(initialdir="C:/", title="Choose your file")
        self.button1.configure(text=str(len(loglists))+" files are selected!")

    def savedir(self):
        savedirectory = filedialog.askdirectory(title="Choose file directory for saving files")
        self.button2.configure(text=str(savedirectory))

    def columnname(self):
        global window
        global window2
        global loglists
        global columnlists
        global listbox

        columnlists = []

        for log in loglists:
            workbook = openpyxl.load_workbook(log)

            sheet1 = workbook.worksheets[0]  # 기존 Data
            for row in sheet1.iter_rows():
                if not row[0].value and not row[1].value:
                    continue
                for i in row:
                    columnlists.append(i.value)
                break

            window2 = Toplevel(window)
            window2.geometry("640x400+230+230")
            window2.title("Choose column")

            scrollbar = Scrollbar(window2)
            scrollbar.pack(side="right", fill="y")

            listbox = Listbox(window2, yscrollcommand=scrollbar.set, selectmode="extended")

            for i in range(len(columnlists)):
                listbox.insert(END, columnlists[i])

            listbox.pack()
            scrollbar.config(command=listbox.yview)

            buttondel = Button(window2, text="Delete selected items", command=lambda: self.deletelists())
            buttondel.pack(pady=20)

            buttonquit = Button(window2, text="Finish", command=lambda: self.endtoplevel(window2))
            buttonquit.pack(pady=20)

    def endtoplevel(self, win):
        win.destroy()
        self.button3.configure(text="Columns are selected")

    def deletelists(self):
        global columnlists
        global listbox

        selection = listbox.curselection()
        if len(selection) == 0:
            return

        for i in range(len(selection)):
            value = listbox.get(selection[0])
            ind = columnlists.index(value)
            del columnlists[ind]
            listbox.delete(selection[0])

    def selectcolumn(self):
        global loglists

        self.button3.configure(text="Please wait")
        self.columnname()

    def startfilter(self, col):
        # find index of selected columns
        for i in range(len(col)):


a = LogFilter()

# dic = {
#         "GroupCycle_AO": 0,
#         "StepNo_AO": 0,
#         # "Gauge2PrsAI": 0,
#         # "Gauge4PrsAI": 0,
#         "T3BP_Set_Pressure": 0,
#         "T3BP_Pressure": 0,
#         # "T3BP_Set_Position": 0,
#         # "T3BP_Position": 0,
#         # "MFC01StPtAO": 0,
#         # "MFC02StPtAO": 0,
#         # "MFC03StPtAO": 0,
#         # "MFC04StPtAO": 0,
#         # "MFC01FlwAI": 0,
#         # "MFC02FlwAI": 0,
#         # "MFC03FlwAI": 0,
#         # "MFC04FlwAI": 0,
#         # "HT06_TmpSpAIO": 0,
#         # "HT06_TmpRdAI": 0,
#         # "HT07_TmpSpAIO": 0,
#         # "HT07_TmpRdAI": 0,
#         # "HT08_TmpSpAIO": 0,
#         # "HT08_TmpRdAI": 0,
#         # "HT06_OutPwCurtAI": 0,
#         # "HT07_OutPwCurtAI": 0,
#         # "HT08_OutPwCurtAI": 0
#         "T3BP_SetGain": 0,
#         "T3BP_SetPhase": 0
# }

# for fileName in files:
#
#     workbook = openpyxl.load_workbook(fileName)
#
#     sheet1 = workbook.worksheets[0]  # 기존 Data
#     sheet2 = workbook.create_sheet()  # Filtering한 Data 넣을 sheet
#     sheet3 = workbook.create_sheet()  # Avg. Std. Max. Min. 정리
#
#     # Data 추출
#     sheet3.append(["Cycle", "Set Pres.", "Avg", "Std", "Max", "Min", "Gain", "Phase"])
#
#     i = 1
#     currentRow = []
#     Gathering = int(20)
#
#     for row in sheet1.iter_rows():
#         if i == 1:
#             j = 0
#             for cell in row:
#                 if cell.value in dic:
#                     dic[cell.value] = j
#                 j += 1
#             for key in dic:
#                 if dic[key] == 0:
#                     continue
#                 currentRow.append(row[dic[key]].value)
#
#             sheet2.append(currentRow)
#             i = 0
#             continue
#
#         currentRow = []
#         for key in dic:
#             if dic[key] == 0:
#                 continue
#             currentRow.append(row[dic[key]].value)
#
#         sheet2.append(currentRow)
#
#     i = 1
#
#     buffer = []
#     a = []
#     dicList = list(dic.keys())
#     GroupNumber = 1
#     Gain = 0
#     Phase = 0
#
#     for row in sheet2.iter_rows():
#
#         if i == 1:
#             i = 0
#             continue
#
#         if int(row[dicList.index("StepNo_AO")].value) % 2 == 1:
#             if not buffer:
#                 continue
#             else:
#                 a = []
#                 a.append(GroupNumber)
#                 a.append(row[dicList.index("T3BP_Set_Pressure")].value)
#                 a.append(numpy.mean(buffer[-Gathering:]))
#                 a.append(numpy.std(buffer[-Gathering:]))
#                 a.append(numpy.max(buffer[-Gathering:]))
#                 a.append(numpy.min(buffer[-Gathering:]))
#                 a.append(Gain)
#                 a.append(Phase)
#                 buffer = []
#                 sheet3.append(a)
#
#                 print(int(row[dicList.index("StepNo_AO")].value))
#                 if int(row[dicList.index("StepNo_AO")].value) == 1:
#                     GroupNumber = row[dicList.index("GroupCycle_AO")].value
#
#         else:
#             buffer.append(row[dicList.index("T3BP_Pressure")].value)
#             Gain = row[dicList.index("T3BP_SetGain")].value
#             Phase = row[dicList.index("T3BP_SetPhase")].value
#
#         if buffer:
#             a = []
#             a.append(GroupNumber)
#             a.append(row[dicList.index("StepNo_AO")].value)
#             a.append(numpy.mean(buffer[-Gathering:]))
#             a.append(numpy.std(buffer[-Gathering:]))
#             a.append(numpy.max(buffer[-Gathering:]))
#             a.append(numpy.min(buffer[-Gathering:]))
#             a.append(Gain)
#             a.append(Phase)
#             buffer = []
#             sheet3.append(a)
#
#     workbook.remove(workbook.worksheets[0])
#     workbook.save(saveDir + "filtered_" + basename(fileName))
