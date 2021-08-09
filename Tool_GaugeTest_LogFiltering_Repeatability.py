import openpyxl
import numpy

# Log에서 남기고 싶은 Parameter

dic = {
        "GroupCycle_AO": 0,
        "StepNo_AO": 0,
        # "Gauge2PrsAI": 0,
        # "Gauge4PrsAI": 0,
        "T3BP_Set_Pressure": 0,
        "T3BP_Pressure": 0,
        "T3BP_Set_Position": 0,
        "T3BP_Position": 0,
        "MFC01StPtAO": 0,
        "MFC02StPtAO": 0,
        "MFC03StPtAO": 0,
        "MFC04StPtAO": 0,
        "MFC01FlwAI": 0,
        "MFC02FlwAI": 0,
        "MFC03FlwAI": 0,
        "MFC04FlwAI": 0,
        # "HT06_TmpSpAIO": 0,
        # "HT06_TmpRdAI": 0,
        # "HT07_TmpSpAIO": 0,
        # "HT07_TmpRdAI": 0,
        # "HT08_TmpSpAIO": 0,
        # "HT08_TmpRdAI": 0,
        # "HT06_OutPwCurtAI": 0,
        # "HT07_OutPwCurtAI": 0,
        # "HT08_OutPwCurtAI": 0
}

#Gathering = input("""계산에 사용할 Data 갯수(숫자만 입력 해주세요): """)
Gathering = int(20)

while True:

    filename = input("""파일명(xlsx 확장자 포함 파일 이름)\nor 종료하려면 Enter: """)

    if not filename:
        break

    workbook = openpyxl.load_workbook(filename)

    sheet1 = workbook.worksheets[0]  # 기존 Data
    sheet2 = workbook.create_sheet()  # Filtering한 Data 넣을 sheet
    sheet3 = workbook.create_sheet()  # Avg. Std. Max. Min. 정리

    # Data 추출
    sheet3.append(["Cycle", "Target Pres.", "Avg", "Std", "Max", "Min"])

    i = 1
    currentRow = []

    for row in sheet1.iter_rows():
        if i == 1:
            j = 0
            for cell in row:
                if cell.value in dic:
                    dic[cell.value] = j
                j += 1
            for key in dic:
                if dic[key] == 0:
                    continue
                currentRow.append(row[dic[key]].value)

            sheet2.append(currentRow)
            i = 0
            continue

        currentRow = []
        for key in dic:
            if dic[key] == 0:
                continue
            currentRow.append(row[dic[key]].value)

        sheet2.append(currentRow)

    i = 1

    buffer = []
    a = []
    dicList = list(dic.keys())
    GroupNumber = 1

    for row in sheet2.iter_rows():

        if i == 1:
            i = 0
            continue

        if int(row[dicList.index("StepNo_AO")].value) % 2 == 1:
            if not buffer:
                continue
            else:
                a = []
                a.append(GroupNumber)
                a.append(row[dicList.index("T3BP_Set_Pressure")].value)
                a.append(numpy.mean(buffer[-Gathering:]))
                a.append(numpy.std(buffer[-Gathering:]))
                a.append(numpy.max(buffer[-Gathering:]))
                a.append(numpy.min(buffer[-Gathering:]))
                buffer = []
                sheet3.append(a)

                if int(row[dicList.index("StepNo_AO")].value) == 1:
                    GroupNumber = row[dicList.index("GroupCycle_AO")].value

        else:
            buffer.append(row[dicList.index("T3BP_Pressure")].value)

    if buffer:
        a = []
        a.append(row[dicList.index("GroupCycle_AO")].value)
        a.append(row[dicList.index("StepNo_AO")].value)
        a.append(numpy.mean(buffer[-Gathering:]))
        a.append(numpy.std(buffer[-Gathering:]))
        a.append(numpy.max(buffer[-Gathering:]))
        a.append(numpy.min(buffer[-Gathering:]))
        buffer = []
        sheet3.append(a)

    workbook.remove(workbook.worksheets[0])
    workbook.save("filtered_" + filename)

print("Complete!")