import openpyxl
import numpy
import tkinter as tk
from tkinter import filedialog
from os.path import basename

dic = {
        "GroupCycle_AO": 0,
        "StepNo_AO": 0,
        # "Gauge2PrsAI": 0,
        # "Gauge4PrsAI": 0,
        "T3BP_Set_Pressure": 0,
        "T3BP_Pressure": 0,
        # "T3BP_Set_Position": 0,
        # "T3BP_Position": 0,
        # "MFC01StPtAO": 0,
        # "MFC02StPtAO": 0,
        # "MFC03StPtAO": 0,
        # "MFC04StPtAO": 0,
        # "MFC01FlwAI": 0,
        # "MFC02FlwAI": 0,
        # "MFC03FlwAI": 0,
        # "MFC04FlwAI": 0,
        # "HT06_TmpSpAIO": 0,
        # "HT06_TmpRdAI": 0,
        # "HT07_TmpSpAIO": 0,
        # "HT07_TmpRdAI": 0,
        # "HT08_TmpSpAIO": 0,
        # "HT08_TmpRdAI": 0,
        # "HT06_OutPwCurtAI": 0,
        # "HT07_OutPwCurtAI": 0,
        # "HT08_OutPwCurtAI": 0
        "T3BP_SetGain": 0,
        "T3BP_SetPhase": 0
}

window = tk.Tk()

window.title("Log Filter")
window.resizable(True, True)

files = filedialog.askopenfilenames(initialdir="C:/", title="Choose your file")

saveDir = filedialog.askdirectory(title="Choose file directory for saving files")

label = tk.Label(window, text=str(len(files))+" file(s) are processing")
label.pack()

for fileName in files:
    # workbook = openpyxl.load_workbook(fileName)
    #
    # sheet1 = workbook.worksheets[0]  # 기존 Data
    # sheet2 = workbook.create_sheet()  # Filtering한 Data 넣을 sheet
    # sheet3 = workbook.create_sheet()  # Avg. Std. Max. Min. 정리
    #
    # i = 1
    # m = 1
    # currentRow = []
    # cycle = 0
    # buffer = []
    # a = [["Step"]]
    #
    # for row in sheet1.iter_rows():
    #     if i == 1:
    #         j = 0
    #         for cell in row:
    #             if cell.value in dic:
    #                 dic[cell.value] = j
    #             j += 1
    #         i = 0
    #         continue
    #     if cycle != row[dic["GroupCycle_AO"]].value:
    #         cycle = row[dic["GroupCycle_AO"]].value
    #         a.append([row[dic["GroupCycle_AO"]].value])
    #
    # i = 1
    # j = 1
    # for row in sheet1.iter_rows():
    #     if i == 1:
    #         i = 0
    #         continue
    #
    #     if j == 1:
    #         if row[dic["StepNo_AO"]].value == 1 and row[dic["GroupCycle_AO"]].value == 2:
    #             j = 0
    #             continue
    #         a[0].append(row[dic["StepNo_AO"]].value)
    #
    #     b = row[dic["GroupCycle_AO"]].value
    #     a[b].append(row[dic["T3BP_Pressure"]].value)
    #
    # j = len(a[0])
    # for i in range(len(a)):
    #     if j > len(a[i]):
    #
    #         j = len(a[i])
    #
    # for p in range(j):
    #     buffer = []
    #     for q in range(len(a)):
    #         buffer.append(a[q][p])
    #     sheet2.append(buffer)
    #
    # workbook.remove(workbook.worksheets[0])

    workbook = openpyxl.load_workbook(fileName)

    sheet1 = workbook.worksheets[0]  # 기존 Data
    sheet2 = workbook.create_sheet()  # Filtering한 Data 넣을 sheet
    sheet3 = workbook.create_sheet()  # Avg. Std. Max. Min. 정리

    # Data 추출
    sheet3.append(["Cycle", "Set Pres.", "Avg", "Std", "Max", "Min", "Gain", "Phase"])

    i = 1
    currentRow = []
    Gathering = int(20)

    for row in sheet1.iter_rows():
        if i == 1:
            j = 0
            for cell in row:
                if cell.value in dic:
                    dic[cell.value] = j
                j += 1
            for key in dic:
                if dic[key] == 0:
                    continue
                currentRow.append(row[dic[key]].value)

            sheet2.append(currentRow)
            i = 0
            continue

        currentRow = []
        for key in dic:
            if dic[key] == 0:
                continue
            currentRow.append(row[dic[key]].value)

        sheet2.append(currentRow)

    i = 1

    buffer = []
    a = []
    dicList = list(dic.keys())
    GroupNumber = 1
    Gain = 0
    Phase = 0

    for row in sheet2.iter_rows():

        if i == 1:
            i = 0
            continue

        if int(row[dicList.index("StepNo_AO")].value) % 2 == 1:
            if not buffer:
                continue
            else:
                a = []
                a.append(GroupNumber)
                a.append(row[dicList.index("T3BP_Set_Pressure")].value)
                a.append(numpy.mean(buffer[-Gathering:]))
                a.append(numpy.std(buffer[-Gathering:]))
                a.append(numpy.max(buffer[-Gathering:]))
                a.append(numpy.min(buffer[-Gathering:]))
                a.append(Gain)
                a.append(Phase)
                buffer = []
                sheet3.append(a)

                print(int(row[dicList.index("StepNo_AO")].value))
                if int(row[dicList.index("StepNo_AO")].value) == 1:
                    GroupNumber = row[dicList.index("GroupCycle_AO")].value

        else:
            buffer.append(row[dicList.index("T3BP_Pressure")].value)
            Gain = row[dicList.index("T3BP_SetGain")].value
            Phase = row[dicList.index("T3BP_SetPhase")].value

    if buffer:
        a = []
        a.append(GroupNumber)
        a.append(row[dicList.index("StepNo_AO")].value)
        a.append(numpy.mean(buffer[-Gathering:]))
        a.append(numpy.std(buffer[-Gathering:]))
        a.append(numpy.max(buffer[-Gathering:]))
        a.append(numpy.min(buffer[-Gathering:]))
        a.append(Gain)
        a.append(Phase)
        buffer = []
        sheet3.append(a)

    workbook.remove(workbook.worksheets[0])
    workbook.save(saveDir + "filtered_" + basename(fileName))

label = tk.Label(window, text="Processing of "+str(len(files))+" file(s) are finished")
label.pack()

window.mainloop()
