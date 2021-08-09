import openpyxl
import numpy as np

# Log에서 남기고 싶은 Parameter

dic = {
        "GroupCycle_AO": 0,
        "StepNo_AO": 0,
        # "Gauge2PrsAI": 0,
        # "Gauge4PrsAI": 0,
        # "T3BP_Set_Pressure": 0,
        "T3BP_Pressure": 0,
        # "T3BP_Set_Position": 0,
        # "T3BP_Position": 0,
        # "MFC01StPtAO": 0,
        # "MFC02StPtAO": 0,
        # "MFC03StPtAO": 0,
        # "MFC04StPtAO": 0,
        # "MFC01FlwAI": 0,
        # "MFC02FlwAI": 0,
        # "MFC03FlwAI": 0,
        # "MFC04FlwAI": 0,
        # "HT06_TmpSpAIO": 0,
        # "HT06_TmpRdAI": 0,
        # "HT07_TmpSpAIO": 0,
        # "HT07_TmpRdAI": 0,
        # "HT08_TmpSpAIO": 0,
        # "HT08_TmpRdAI": 0,
        # "HT06_OutPwCurtAI": 0,
        # "HT07_OutPwCurtAI": 0,
        # "HT08_OutPwCurtAI": 0
        # "T3BP_SetGain": 0,
        # "T3BP_SetPhase": 0
}

filename = input("""파일명(xlsx 확장자 포함 파일 이름): """)

workbook = openpyxl.load_workbook(filename)

sheet1 = workbook.worksheets[0]  # 기존 Data
sheet2 = workbook.create_sheet()  # Filtering한 Data 넣을 sheet
sheet3 = workbook.create_sheet()  # Avg. Std. Max. Min. 정리

i = 1
m = 1
currentRow = []
cycle = 0
buffer = []
a = [["Step"]]

for row in sheet1.iter_rows():
    if i == 1:
        j = 0
        for cell in row:
            if cell.value in dic:
                dic[cell.value] = j
            j += 1
        i = 0
        continue
    if cycle != row[dic["GroupCycle_AO"]].value:
        cycle = row[dic["GroupCycle_AO"]].value
        a.append([row[dic["GroupCycle_AO"]].value])

i = 1
j = 1
for row in sheet1.iter_rows():
    if i == 1:
        i = 0
        continue

    if j == 1:
        if row[dic["StepNo_AO"]].value == 1 and row[dic["GroupCycle_AO"]].value == 2:
            j = 0
            continue
        a[0].append(row[dic["StepNo_AO"]].value)

    b = row[dic["GroupCycle_AO"]].value
    a[b].append(row[dic["T3BP_Pressure"]].value)

j = len(a[0])
for i in range(len(a)):
    if j > len(a[i]):
        j = len(a[i])

for p in range(j):
    buffer = []
    for q in range(len(a)):
        buffer.append(a[q][p])
    sheet2.append(buffer)

workbook.remove(workbook.worksheets[0])
workbook.save("filtered_11" + filename)
